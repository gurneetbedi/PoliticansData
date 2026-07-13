"""
Convert an ECI Statistical Report "Detailed Results" Excel file into the
same JSON schema that `scripts/fetch_eci_results.py` produces.

Use this for elections whose live results portal has been retired
(e.g. Odisha / Andhra Pradesh / Sikkim / Arunachal Pradesh June 2024).
Download the "10-Detailed-Results.xlsx" from
    https://www.eci.gov.in/statistical-report/ae/<year>/<num>
and pass it here.

Usage:
    python scripts/parse_eci_statistical_report.py \\
        --excel data/10-Detailed-Results.xlsx \\
        --state "Odisha" \\
        --year 2024 \\
        --state-code S18 \\
        --out data/eci/results/odisha_2024_eci_results.json

The downstream loader `load_eci_results.py` doesn't care where the JSON
came from — it just wants the schema. That means the rest of the
ingestion pipeline (build allowlist → OCR → apply → load results) is
unchanged.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path


# Column indices — from row 3 (the header row) of the standard ECI
# "Detailed Results" Excel template.
COL_STATE   = 0
COL_AC_NO   = 1
COL_AC_NAME = 2
COL_CAND    = 3
COL_PARTY   = 7
COL_SYMBOL  = 8
COL_EVM     = 9
COL_POSTAL  = 10
COL_TOTAL   = 11
COL_PCT     = 12


def _norm_name(raw: str | None) -> str:
    """Strip the leading "N " rank number ECI prefixes and uppercase."""
    if raw is None:
        return ""
    s = str(raw).strip()
    # Candidate cells are formatted like "1 Barsha Singh Bariha" — the
    # leading number is that candidate's *listing* order (usually
    # ballot position), not their result rank. Strip it.
    s = re.sub(r"^\s*\d+\s+", "", s)
    return s.strip().upper()


def _to_int(v) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return None


def _to_float(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_excel(path: Path) -> list[dict]:
    """Read the Excel and return a list of constituency dicts matching the
    schema of `fetch_eci_results.py` output.
    """
    try:
        import openpyxl
    except ImportError:
        sys.exit("pip install openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    constituencies: dict[int, dict] = {}

    header_seen = False
    for row in ws.iter_rows(values_only=True):
        # Skip until we pass the "STATE/UT NAME" header row.
        if not header_seen:
            if row and str(row[0] or "").strip().upper() == "STATE/UT NAME":
                header_seen = True
            continue

        # Boundary rows say "TURN OUT" in col 0 — skip them.
        first_cell = str(row[0] or "").strip().upper()
        if first_cell in ("", "TURN OUT", "GRAND TOTAL"):
            continue

        try:
            ac_no = _to_int(row[COL_AC_NO])
            if ac_no is None:
                continue
            ac_name = str(row[COL_AC_NAME] or "").strip()
            candidate = _norm_name(row[COL_CAND])
            party = str(row[COL_PARTY] or "").strip().upper()
            evm = _to_int(row[COL_EVM])
            postal = _to_int(row[COL_POSTAL])
            total = _to_int(row[COL_TOTAL])
            pct = _to_float(row[COL_PCT])
        except IndexError:
            continue

        if not candidate:
            continue

        # NOTA is stored as its own candidate row — keep it (matches the
        # 2026 fetcher's output, and downstream loaders already skip
        # matching NOTA to any real politician).
        if candidate == "NOTA":
            party = "NOTA"

        c = constituencies.setdefault(ac_no, {
            "number": ac_no,
            "name": ac_name.upper(),
            "candidates": [],
        })
        c["candidates"].append({
            "name":         candidate,
            "party":        party,
            "evm_votes":    evm,
            "postal_votes": postal,
            "total_votes":  total,
            "vote_pct":     pct,
        })

    # Rank candidates within each constituency by total_votes desc,
    # assign rank + won flag. Winner = highest total_votes.
    out = []
    for num in sorted(constituencies):
        c = constituencies[num]
        c["candidates"].sort(key=lambda x: -(x["total_votes"] or 0))
        for i, cand in enumerate(c["candidates"], 1):
            cand["rank"] = i
            cand["won"] = (i == 1)
        out.append(c)
    return out


def main():
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--excel", required=True,
                    help="Path to the '10-Detailed-Results.xlsx' downloaded "
                         "from eci.gov.in/statistical-report/")
    ap.add_argument("--state", required=True,
                    help='State name in TitleCase, e.g. "Odisha"')
    ap.add_argument("--year", type=int, required=True)
    ap.add_argument("--state-code", required=True,
                    help="ECI state code (S18 = Odisha, S01 = AP, etc.)")
    ap.add_argument("--out", required=True,
                    help="Output JSON path")
    args = ap.parse_args()

    # Normalize --state to canonical TitleCase (matches the other scripts
    # in the pipeline). Multi-word states have both spaced and no-space
    # forms mapped so `andhrapradesh` → "Andhra Pradesh" transparently.
    if args.state:
        _SPECIAL = {
            "jammu and kashmir": "Jammu and Kashmir",
            "jammuandkashmir":   "Jammu and Kashmir",
            "jk":                "Jammu and Kashmir",
            "andhra pradesh":    "Andhra Pradesh",
            "andhrapradesh":     "Andhra Pradesh",
            "arunachal pradesh": "Arunachal Pradesh",
            "arunachalpradesh":  "Arunachal Pradesh",
            "himachal pradesh":  "Himachal Pradesh",
            "himachalpradesh":   "Himachal Pradesh",
            "madhya pradesh":    "Madhya Pradesh",
            "madhyapradesh":     "Madhya Pradesh",
            "tamil nadu":        "Tamil Nadu",
            "tamilnadu":         "Tamil Nadu",
            "uttar pradesh":     "Uttar Pradesh",
            "uttarpradesh":      "Uttar Pradesh",
            "west bengal":       "West Bengal",
            "westbengal":        "West Bengal",
        }
        lc = args.state.strip().lower()
        args.state = _SPECIAL.get(lc, args.state.strip().title())

    excel_path = Path(args.excel)
    if not excel_path.exists():
        sys.exit(f"Excel file not found: {excel_path}")

    print(f"→ Parsing {excel_path.name} ...", file=sys.stderr)
    constituencies = parse_excel(excel_path)
    total_cands = sum(len(c["candidates"]) for c in constituencies)
    print(f"  {len(constituencies)} constituencies, {total_cands} candidates",
          file=sys.stderr)

    if not constituencies:
        sys.exit("No constituencies parsed — check the file structure.")

    payload = {
        "state":         args.state,
        "year":          args.year,
        "state_code":    args.state_code,
        "source":        f"ECI Statistical Report - Detailed Results ({excel_path.name})",
        "assembly_size": len(constituencies),
        "constituencies": constituencies,
    }

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False))
    print(f"→ Wrote {out_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
