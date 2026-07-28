"""RBI Handbook of Statistics on Indian States → GPI indicator ingester (Tier A).

The RBI Handbook is published annually as a set of ~130 XLSX tables, each
covering one state-level statistical series. Latest edition observed:
11 Dec 2025 (URLs in RBI_TABLES below encode that publication date; when RBI
releases a new edition, refresh the URLs and re-run --download).

This script covers 3 indicators from the RBI Handbook:

    E01   Real GSDP growth rate            (Table 22: GSDP Constant Prices)
    E02   Per-capita NSDP (current prices) (Table 19)
    H01   Infant Mortality Rate            (Table 4)

Fiscal indicators F01-F06 come from the separate "State Finances: A Study
of Budgets" publication — that ingester will live in a companion script.

Usage:
    # One-shot: download + ingest all supported indicators for all states
    python scripts/gpi_ingest_rbi_handbook.py --download --ingest

    # Just download (skip if files already exist)
    python scripts/gpi_ingest_rbi_handbook.py --download

    # Ingest only (reads previously-downloaded files)
    python scripts/gpi_ingest_rbi_handbook.py --ingest

    # Preview without writing to DB
    python scripts/gpi_ingest_rbi_handbook.py --ingest --dry-run

    # Punjab pilot only
    python scripts/gpi_ingest_rbi_handbook.py --ingest --state PB

    # One indicator (useful for debugging XLSX parsing)
    python scripts/gpi_ingest_rbi_handbook.py --ingest --indicators E01

Files land in:  data/gpi/rbi/handbook/{edition}/table_{N}.xlsx
"""
from __future__ import annotations
import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

# ═══════════════════════════════════════════════════════════════════════════
# RBI publication config — one dict per indicator we source from the Handbook.
# Update `url` values when RBI publishes a new edition (URLs contain a date
# stamp and hash so they change per release).
# ═══════════════════════════════════════════════════════════════════════════
EDITION = "2024-25"   # latest observed publication date suffix
EDITION_DIR = ROOT / "data" / "gpi" / "rbi" / "handbook" / EDITION

RBI_TABLES = {
    "E01": {
        "table_num": 22,
        "table_name": "Gross State Domestic Product (Constant Prices)",
        # `compute="yoy_growth"` → the XLSX has absolute GSDP values; we
        # compute year-on-year % growth in Python. Result is what E01 measures.
        "compute": "yoy_growth",
        "unit": "% annual",
        "url": "https://rbidocs.rbi.org.in/rdocs/Publications/DOCs/22T_11122025E6BC0CB35180406EAB6E0D49DE51C8E8.XLSX",
        "source_document": "RBI Handbook of Statistics on Indian States 2024-25, Table 22",
    },
    "E02": {
        "table_num": 19,
        "table_name": "Per Capita Net State Domestic Product (Current Prices)",
        "compute": "direct",   # XLSX already holds per-capita values in INR
        "unit": "INR/capita",
        "url": "https://rbidocs.rbi.org.in/rdocs/Publications/DOCs/19T_11122025B8CC230E4A34431999B4D6A107707BCA.XLSX",
        "source_document": "RBI Handbook of Statistics on Indian States 2024-25, Table 19",
    },
    "H01": {
        "table_num": 4,
        "table_name": "State-wise Infant Mortality Rate",
        "compute": "direct",   # values are IMR per 1000 live births
        "unit": "per 1000",
        "url": "https://rbidocs.rbi.org.in/rdocs/Publications/DOCs/4T_11122025025F203A250E46CAB963946C776ADBAF.XLSX",
        "source_document": "RBI Handbook of Statistics on Indian States 2024-25, Table 4",
    },
}


# ═══════════════════════════════════════════════════════════════════════════
# State-name normalization — RBI uses various forms across editions.
# Maps normalized (lower-cased, punctuation-stripped) source names to the
# canonical names in our `states` table.
# ═══════════════════════════════════════════════════════════════════════════
STATE_ALIASES = {
    "orissa": "Odisha",
    "odisha": "Odisha",
    "jammu & kashmir": "Jammu and Kashmir",
    "jammu and kashmir": "Jammu and Kashmir",
    "jammu & kashmir (ut)": "Jammu and Kashmir",
    "jammu and kashmir (ut)": "Jammu and Kashmir",
    "j&k": "Jammu and Kashmir",
    "j & k": "Jammu and Kashmir",
    "nct of delhi": "Delhi",
    "delhi (ut)": "Delhi",
    "delhi": "Delhi",
    "chattisgarh": "Chhattisgarh",
    "chhattisgarh": "Chhattisgarh",
    "uttaranchal": "Uttarakhand",
    "uttarakhand": "Uttarakhand",
    "pondicherry": "Puducherry",
    "puducherry": "Puducherry",
    # Rows we should NOT insert as states (footer totals, headings, notes)
    "all-india": None,
    "all india": None,
    "india": None,
    "average": None,
    "total": None,
}


def normalize_state_name(raw: str) -> str | None:
    """Map RBI's state string to our canonical DB name. None means 'skip'."""
    if not raw:
        return None
    n = str(raw).strip().lower()
    n = re.sub(r"[.*#$@]", "", n)
    n = re.sub(r"\s+", " ", n).strip()
    if n in STATE_ALIASES:
        return STATE_ALIASES[n]
    # Title-case the input as a fallback so "west bengal" → "West Bengal"
    return " ".join(w.capitalize() for w in n.split())


def parse_year_from_header(cell) -> int | None:
    """RBI year headers: '2015-16', '2015-2016', 2016 → 2016 (fiscal year end).

    Returns None if the cell doesn't look like a year (e.g. header labels,
    unit annotations, blanks, or truly numeric 4-digit calendar years like
    2020 which we treat as-is).
    """
    if cell is None:
        return None
    s = str(cell).strip()
    if not s or s.lower() in ("state", "state/ut", "sr.no", "s.no.", "s no"):
        return None
    m = re.match(r"^(\d{4})[-/]\d{2,4}$", s)     # "2015-16" or "2015-2016"
    if m:
        return int(m.group(1)) + 1
    m = re.match(r"^(\d{4})$", s)                 # plain 4-digit year
    if m:
        y = int(m.group(1))
        if 1990 <= y <= 2100:
            return y
    return None


# ═══════════════════════════════════════════════════════════════════════════
# Download
# ═══════════════════════════════════════════════════════════════════════════
def _is_valid_xlsx(path: Path) -> bool:
    """Real XLSX = ZIP archive; first 4 bytes are PK\\x03\\x04.
    HTML error pages start with '<' — anything not starting with PK is not XLSX."""
    try:
        with path.open("rb") as f:
            head = f.read(4)
        return head[:2] == b"PK"
    except Exception:
        return False


# Full browser-like header set — RBI's CDN blocks unadorned requests and
# returns an HTML "please use browser" page instead of the XLSX. Referer
# spoofs a click from the Handbook landing page.
_RBI_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/126.0.0.0 Safari/537.36"
    ),
    "Accept": (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
        "application/vnd.ms-excel,application/octet-stream,*/*;q=0.8"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.rbi.org.in/scripts/AnnualPublications.aspx?head=Handbook%20of%20Statistics%20on%20Indian%20States",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "same-site",
}


def download_table(indicator_code: str, config: dict, out_dir: Path) -> Path:
    """Fetch one RBI Handbook XLSX and save to disk. Validates that we got
    a real XLSX (ZIP magic bytes); if RBI returned HTML instead, deletes the
    file and raises with a clear error pointing at the manual-download URL."""
    import requests
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"table_{config['table_num']}_{indicator_code}.xlsx"

    # Purge any stale HTML-in-XLSX-clothing from previous runs
    if out_path.exists() and not _is_valid_xlsx(out_path):
        out_path.unlink()

    if out_path.exists():
        size = out_path.stat().st_size
        if size > 1000 and _is_valid_xlsx(out_path):
            print(f"  ✓ {indicator_code}  already have {out_path.name} ({size // 1024}KB)")
            return out_path
        out_path.unlink()

    print(f"  ↓ {indicator_code}  fetching Table {config['table_num']} ...")
    # Use a session so cookies (if any) get set by the landing page hit.
    with requests.Session() as sess:
        sess.headers.update(_RBI_HEADERS)
        # Prime session with a hit to the landing page — some RBI CDN
        # paths refuse direct-file requests without a prior page view.
        try:
            sess.get(_RBI_HEADERS["Referer"], timeout=30)
        except Exception:
            pass  # priming is best-effort
        resp = sess.get(config["url"], timeout=60)
        resp.raise_for_status()
        out_path.write_bytes(resp.content)

    # Validate
    if not _is_valid_xlsx(out_path):
        # Show first 200 bytes so user can see what came back
        with out_path.open("rb") as f:
            head = f.read(200)
        out_path.unlink()
        raise RuntimeError(
            f"\n  ✗ {indicator_code}: RBI returned non-XLSX content "
            f"(likely HTML block page).\n"
            f"    Preview of response: {head[:120]!r}\n"
            f"\n"
            f"    Workaround — download in a browser and drop into place:\n"
            f"      URL:    {config['url']}\n"
            f"      Save as: {out_path}\n"
            f"    Then re-run with --ingest (skip --download).\n"
        )

    print(f"  ✓ {indicator_code}  saved {out_path.name} ({len(resp.content) // 1024}KB)")
    return out_path


# ═══════════════════════════════════════════════════════════════════════════
# XLSX parsing
# ═══════════════════════════════════════════════════════════════════════════
def parse_rbi_table(xlsx_path: Path) -> dict[str, dict[int, float]]:
    """Parse one RBI Handbook XLSX. Returns {state_name: {year: value}}.

    RBI format (empirically stable across 2020-2025 editions):
      Row 1-2 : Title (merged, plain text)
      Row 3-4 : Column headers — first cell is 'State' or 'Sl.No / State',
                remaining cells are years like '2015-16', '2016-17', ...
      Row 5+  : Data rows — Column A optionally a serial number, Column B
                (or A) the state name, following columns the yearly values.
      Bottom  : 'All-India' or 'Total' row (skipped via STATE_ALIASES).

    Some tables number-column-first (Col A = 1..28, Col B = state); some
    put state in Col A directly. We auto-detect by scanning for the first
    non-empty text cell in each data row.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit(
            "openpyxl not installed. Install with:\n"
            "    pip install openpyxl\n"
        )

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # ── Locate the header row (first row where multiple cells parse as years) ─
    header_row_idx = None
    year_by_col = {}
    for row_idx in range(1, min(15, ws.max_row + 1)):
        col_years = {}
        for col_idx in range(1, ws.max_column + 1):
            y = parse_year_from_header(ws.cell(row=row_idx, column=col_idx).value)
            if y:
                col_years[col_idx] = y
        if len(col_years) >= 3:            # need at least 3 year columns to trust it
            header_row_idx = row_idx
            year_by_col = col_years
            break

    if not header_row_idx:
        raise ValueError(f"Could not locate year-header row in {xlsx_path.name}")

    # ── Scan data rows below the header ────────────────────────────────────
    out: dict[str, dict[int, float]] = {}
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        # Find state name: first non-numeric text cell in the row
        state_raw = None
        for col_idx in range(1, ws.max_column + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            # Serial numbers (1., 2., "1)") — skip
            if re.match(r"^\d+[.\)]?$", s):
                continue
            state_raw = s
            break

        if not state_raw:
            continue
        state = normalize_state_name(state_raw)
        if state is None:                # explicit skip (All-India, Total, etc.)
            continue

        # Extract values from the identified year columns
        years_for_state = {}
        for col_idx, year in year_by_col.items():
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            if isinstance(v, str):
                v_str = v.strip()
                if v_str in ("", "-", "NA", "N.A.", "n.a.", "—"):
                    continue
                try:
                    v = float(v_str.replace(",", ""))
                except ValueError:
                    continue
            if isinstance(v, (int, float)):
                years_for_state[year] = float(v)

        if years_for_state:
            # If a state appears twice in the sheet (rare), keep the last row's values
            out[state] = {**out.get(state, {}), **years_for_state}

    return out


def compute_yoy_growth(state_values: dict[int, float]) -> dict[int, float]:
    """Convert {year: absolute value} → {year: YoY growth %}."""
    years = sorted(state_values.keys())
    growth = {}
    for i in range(1, len(years)):
        prev, curr = years[i-1], years[i]
        # Only compute for consecutive years
        if curr - prev != 1:
            continue
        v_prev, v_curr = state_values[prev], state_values[curr]
        if v_prev == 0:
            continue
        growth[curr] = round(100.0 * (v_curr - v_prev) / v_prev, 2)
    return growth


# ═══════════════════════════════════════════════════════════════════════════
# DB insert
# ═══════════════════════════════════════════════════════════════════════════
def ingest_indicator(session, indicator_code: str, config: dict, xlsx_path: Path,
                       state_filter: str | None, dry_run: bool):
    from app.gpi_models import GpiIndicator, GpiIndicatorValue
    from app.models import State

    indicator = session.query(GpiIndicator).filter_by(code=indicator_code).one_or_none()
    if not indicator:
        print(f"  ✗ Indicator {indicator_code} not found in DB — run gpi_seed.py first")
        return {"inserted": 0, "updated": 0, "skipped": 0}

    states_by_name = {s.name: s for s in session.query(State).all()}

    parsed = parse_rbi_table(xlsx_path)
    print(f"  📄 Parsed {len(parsed)} states from {xlsx_path.name}")

    counts = {"inserted": 0, "updated": 0, "skipped": 0}
    for state_name, year_values in parsed.items():
        st = states_by_name.get(state_name)
        if not st:
            counts["skipped"] += 1
            continue
        if state_filter and st.code != state_filter:
            continue

        # Apply computation
        if config["compute"] == "yoy_growth":
            final = compute_yoy_growth(year_values)
        else:  # "direct"
            final = year_values

        for year, value in final.items():
            # Only ingest data for our Phase 1 window (2018-2026)
            if year < 2018 or year > 2026:
                continue

            existing = session.query(GpiIndicatorValue).filter_by(
                indicator_id=indicator.id, state_id=st.id, fiscal_year=year
            ).one_or_none()

            payload = {
                "raw_value": value,
                "source_url": config["url"],
                "source_document": config["source_document"],
                "extraction_method": "scraped",
                "staleness": "current",
                "extracted_at": datetime.utcnow(),
            }

            if existing:
                changed = False
                for k, v in payload.items():
                    if getattr(existing, k) != v:
                        setattr(existing, k, v)
                        changed = True
                if changed:
                    existing.normalized_value = None
                    existing.national_rank = None
                    counts["updated"] += 1
            else:
                session.add(GpiIndicatorValue(
                    indicator_id=indicator.id,
                    state_id=st.id,
                    fiscal_year=year,
                    **payload,
                ))
                counts["inserted"] += 1

    if not dry_run:
        session.commit()

    return counts


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                  formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--download", action="store_true",
                    help="Download RBI XLSX tables to data/gpi/rbi/handbook/")
    ap.add_argument("--ingest", action="store_true",
                    help="Parse downloaded files and insert into DB")
    ap.add_argument("--dry-run", action="store_true",
                    help="Ingest mode: preview without writing")
    ap.add_argument("--state", default=None,
                    help="2-letter state code (e.g. PB) — only ingest this state")
    ap.add_argument("--indicators", default=None,
                    help="Comma-separated indicator codes (e.g. E01,E02). Default: all in RBI_TABLES.")
    args = ap.parse_args()

    if not (args.download or args.ingest):
        ap.error("Nothing to do — pass --download and/or --ingest")

    targets = list(RBI_TABLES.keys())
    if args.indicators:
        want = set(args.indicators.split(","))
        targets = [t for t in targets if t in want]
    if not targets:
        ap.error("No matching indicators found in RBI_TABLES")

    print(f"Indicators: {targets}")
    print(f"Edition dir: {EDITION_DIR.relative_to(ROOT)}")
    print()

    # ── DOWNLOAD ────────────────────────────────────────────────────────────
    xlsx_paths = {}
    if args.download:
        print("═══ Download ═══")
        for code in targets:
            xlsx_paths[code] = download_table(code, RBI_TABLES[code], EDITION_DIR)
        print()

    # If not downloading, still need the paths for ingest
    if args.ingest and not args.download:
        for code in targets:
            p = EDITION_DIR / f"table_{RBI_TABLES[code]['table_num']}_{code}.xlsx"
            if not p.exists():
                raise SystemExit(f"Missing {p} — run with --download first")
            if not _is_valid_xlsx(p):
                raise SystemExit(
                    f"\n{p} is not a valid XLSX (likely HTML from a bot-blocked "
                    f"RBI request).\n"
                    f"Delete it and either:\n"
                    f"  • re-run with --download (uses browser-like headers), OR\n"
                    f"  • download manually via browser: {RBI_TABLES[code]['url']}\n"
                    f"    and save as {p}\n"
                )
            xlsx_paths[code] = p

    # ── INGEST ──────────────────────────────────────────────────────────────
    if args.ingest:
        # Import here so --download alone works without SQLAlchemy install
        from app.database import SessionLocal
        from app import models  # ensures gpi_models loaded

        session = SessionLocal()
        try:
            print("═══ Ingest ═══")
            total = {"inserted": 0, "updated": 0, "skipped": 0}
            for code in targets:
                print(f"\n{code} — {RBI_TABLES[code]['table_name']}")
                counts = ingest_indicator(session, code, RBI_TABLES[code], xlsx_paths[code],
                                            state_filter=args.state, dry_run=args.dry_run)
                for k in total:
                    total[k] += counts[k]
                print(f"  ↳ ins={counts['inserted']}  upd={counts['updated']}  "
                      f"skip_unknown_state={counts['skipped']}")
                if args.dry_run:
                    session.rollback()

            print()
            print("═══════════════════ Ingest Summary ═══════════════════")
            print(f"  Inserted: {total['inserted']}")
            print(f"  Updated:  {total['updated']}")
            print(f"  Skipped (state-name unresolved): {total['skipped']}")
            if args.dry_run:
                print("  (DRY RUN — no writes)")
            print()
            print("Next: python scripts/gpi_compute_scores.py")
        finally:
            session.close()


if __name__ == "__main__":
    main()
