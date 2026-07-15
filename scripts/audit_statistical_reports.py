"""
Sanity-check every ECI Statistical Report Excel in data/Results/ before
running the parse pipeline against them.

For each Excel it:
  1. Reads the first data row's STATE/UT NAME value
  2. Guesses the intended state from the filename
  3. Counts distinct AC numbers (should match assembly size)
  4. Flags mismatches (wrong state, off-count, mislabeled file)

Run:
    python scripts/audit_statistical_reports.py
"""
from __future__ import annotations

import re
import sys
from pathlib import Path


EXPECTED_SEATS = {
    "andhra pradesh":     175, "arunachal pradesh":  60, "assam":         126,
    "bihar":              243, "chhattisgarh":       90, "delhi":          70,
    "goa":                 40, "gujarat":           182, "haryana":        90,
    "himachal pradesh":    68, "jammu and kashmir":  90, "jharkhand":      81,
    "karnataka":          224, "kerala":            140, "madhya pradesh":230,
    "maharashtra":        288, "manipur":            60, "meghalaya":      60,
    "mizoram":             40, "nagaland":           60, "odisha":        147,
    "puducherry":          30, "punjab":            117, "rajasthan":     200,
    "sikkim":              32, "tamil nadu":        234, "telangana":     119,
    "tripura":             60, "uttar pradesh":     403, "uttarakhand":    70,
    "west bengal":        294,
}


def guess_state_from_filename(name: str) -> str:
    """Extract the state name from filenames like:
        10-Detailed-Results-Andhra_Pradesh_2024.xlsx
        10-Detailed-Results_Bihar_2025.xlsx
        10-Detailed-Results_ArunachalPardesh_2024.xlsx  (note typo)
        10-Detailed-Results_J&K_2024.xlsx
    Handles both '-' and '_' separators after "Results", the "Pardesh"
    typo, and short-hand abbreviations like J&K.
    """
    stem = Path(name).stem
    # Strip prefix — support BOTH separators after "Results"
    stem = re.sub(r"^\d+-Detailed-Results[-_]", "", stem, flags=re.IGNORECASE)
    # Strip trailing year
    stem = re.sub(r"[_\s-]?\d{4}$", "", stem)
    # Fix common typos
    stem = stem.replace("Pardesh", "Pradesh")
    # Insert space at CamelCase boundaries: MadhyaPradesh → Madhya Pradesh
    stem = re.sub(r"([a-z])([A-Z])", r"\1 \2", stem)
    lc = stem.replace("_", " ").strip().lower()
    # Filename-only aliases
    ALIASES = {
        "j&k": "jammu and kashmir",
        "jk":  "jammu and kashmir",
    }
    return ALIASES.get(lc, lc)


def guess_year_from_filename(name: str) -> str:
    m = re.search(r"(20\d{2})", name)
    return m.group(1) if m else "?"


def inspect(path: Path) -> dict:
    try:
        import openpyxl
    except ImportError:
        sys.exit("pip install openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    # Find header row + first data row
    state_val = None
    ac_numbers = set()
    header_seen = False
    for row in ws.iter_rows(max_row=5000, values_only=True):
        if not header_seen:
            if row and str(row[0] or "").strip().upper() == "STATE/UT NAME":
                header_seen = True
            continue
        c0 = str(row[0] or "").strip()
        if not c0 or c0.upper() in ("TURN OUT", "GRAND TOTAL", "DISCLAIMER"):
            continue
        if state_val is None:
            state_val = c0
        try:
            n = int(row[1]) if row[1] is not None else None
            if n is not None:
                ac_numbers.add(n)
        except (TypeError, ValueError):
            pass
    return {
        "file_state": state_val or "",
        "constituencies": len(ac_numbers),
    }


def main():
    root = Path(__file__).resolve().parent.parent
    results_dir = root / "data" / "Results"
    if not results_dir.exists():
        sys.exit(f"Directory not found: {results_dir}")

    excels = sorted(results_dir.glob("*.xlsx"))
    if not excels:
        sys.exit(f"No .xlsx files in {results_dir}")

    print(f"{'Filename':55s}  {'Guess':22s}  {'InFile':22s}  {'Seats':>6s}  Status")
    print("-" * 130)

    issues = []
    for path in excels:
        guess = guess_state_from_filename(path.name)
        year = guess_year_from_filename(path.name)
        try:
            info = inspect(path)
        except Exception as e:
            print(f"{path.name:55s}  {guess:22s}  ERROR: {e}")
            issues.append((path.name, f"parse error: {e}"))
            continue

        in_file = (info["file_state"] or "").strip().lower()
        expected = EXPECTED_SEATS.get(guess, None)

        # Normalize both sides for comparison — collapse punctuation +
        # "and"/"&", strip "nct of", trim whitespace.
        def _norm_st(s: str) -> str:
            s = s.lower()
            s = s.replace("&", "and")
            s = re.sub(r"^\s*nct\s+of\s+", "", s)
            s = re.sub(r"[^a-z]+", "", s)
            return s

        state_ok = _norm_st(in_file) == _norm_st(guess) or (
            _norm_st(guess) and _norm_st(guess) in _norm_st(in_file)
        ) or (
            _norm_st(in_file) and _norm_st(in_file) in _norm_st(guess)
        )
        count_ok = expected is not None and info["constituencies"] == expected
        count_close = expected is not None and abs(info["constituencies"] - expected) <= 2

        if not state_ok:
            status = f"⚠ STATE MISMATCH — file says {info['file_state']!r}, name says {guess!r}"
            issues.append((path.name, status))
        elif count_ok:
            status = "✓ OK"
        elif count_close:
            status = f"⚠ close: expected {expected}, got {info['constituencies']}"
        else:
            status = f"✗ count off: expected {expected}, got {info['constituencies']}"
            issues.append((path.name, status))

        print(f"{path.name:55s}  {guess[:22]:22s}  "
              f"{info['file_state'][:22]:22s}  "
              f"{info['constituencies']:>6d}  {status}")

    print("-" * 130)
    if issues:
        print(f"\n⚠ {len(issues)} file(s) need attention:")
        for name, msg in issues:
            print(f"  {name}: {msg}")
    else:
        print(f"\n✓ All {len(excels)} files verified. Safe to run parse pipeline.")


if __name__ == "__main__":
    main()
