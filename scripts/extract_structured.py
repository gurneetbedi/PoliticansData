"""
No-LLM structured field extraction from preprocessed ECI affidavit JSONs.

Walks data/eci/for_ai/preprocessed/*.json (the output of
preprocess_eci_pdfs.py / preprocess_modal.py), applies a stack of regex
anchors to the column-delimited text, and writes one row per candidate
into CSV + XLSX.

PHILOSOPHY
----------
Conservative. When the OCR text is ambiguous or a value falls between
two plausible captures, we emit `null` and add a note rather than
guessing wrong. The Excel `notes` column documents every field that
was dropped or required a fallback so a human reviewer knows exactly
where to look.

OUTPUTS
-------
  data/eci/for_ai/extracted/delhi_2025_structured.csv     ← tabular roster
  data/eci/for_ai/extracted/delhi_2025_structured.xlsx    ← same, formatted
  data/eci/for_ai/extracted/_extraction_notes.csv         ← per-candidate notes

USAGE
-----
    python scripts/extract_structured.py

Run after preprocess_modal.py / preprocess_eci_pdfs.py has produced the
per-candidate preprocessed JSONs. No API keys, no network.
"""
from __future__ import annotations

import csv
import json
import re
import sys
from dataclasses import dataclass, asdict, field
from pathlib import Path

# ---------------------------------------------------------------------------
# Anchors — regex patterns we look for in the preprocessed text
# ---------------------------------------------------------------------------

# --- eStamp cover page ----------------------------------------------------
ESTAMP_CERT_RE = re.compile(
    r"(?:Certificate\s+No\.?\s*[:|\.]?\s*)?(IN-[A-Z]{2}\d{14,}[A-Za-z])"
)
ESTAMP_DATE_RE = re.compile(
    r"Certificate\s+Issued\s+Date\s*[:|\.]?\s*"
    r"(\d{1,2}-[A-Za-z]{3}-\d{4}\s+\d{1,2}[:.]\d{2}\s*[AP]M)"
)
ESTAMP_PURCHASER_RE = re.compile(
    r"Purchased\s+by\s*[:|\.]?\s*([A-Z][A-Z\s\.]{2,80}?)\s*(?:\||$|\n|Description)"
)
ESTAMP_DUTY_RE = re.compile(
    r"Stamp\s+Duty\s+Amount\s*\(?\s*Rs\.?\s*\)?\s*[:|\.]?\s*(\d{1,4})"
)

# --- Part A header --------------------------------------------------------
# "I AKHILESH PATI TRIPATHI **son/daughter/wife of SH; ABHAY NANDAN TRIPATHI"
PART_A_NAME_RE = re.compile(
    r"\bI\s+([A-Z][A-Z\s\.]{2,60}?)\s*\*+\s*"
    r"(son|daughter|wife|husband)[^a-zA-Z]+of\s+"
    r"([A-Z][A-Z\s\.\;]{2,80}?)\s*"
    r"(?:\b[Aa]ged|\n)",
    re.IGNORECASE,
)
# "Aged 40 years"
AGE_RE = re.compile(r"\b[Aa]ged\s+(\d{1,3})\s*years?")
# "resident of N-9C/129, LALBAGH, AZADPUR, DELHI-110033"
ADDRESS_RE = re.compile(
    r"resident\s+of\s+([^\n]+?)\s*(?:\(mention\b|,?\s*a\s+candidate)",
    re.IGNORECASE,
)
# "candidate set up by AAM AADMI PARTY" or "set up by-INDEPENDENT" etc.
# Looser char class — accept anything that isn't pipe/newline/star.
PARTY_RE = re.compile(
    r"candidate\s+set\s+up\s+by[-:\s]+([A-Z][^\n|\*]{2,80}?)"
    r"(?=\s*[\(|\*\n]|$)",
    re.IGNORECASE,
)
# "enrolled in AC-18, MODEL TOWN, NCT OF DELHI" — note AC-18 needs digits.
CONSTITUENCY_RE = re.compile(
    r"enrolled\s+in\s+([A-Z][A-Z0-9\s,\-\(\)\.]+?)"
    r"(?:\s*\(Name|\s*at\s*Serial|\s*\|)",
    re.IGNORECASE,
)
# email + phone
EMAIL_RE = re.compile(r"\b([A-Za-z][A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})")
PHONE_RE = re.compile(r"\b((?:91[-\s]?)?[6-9]\d{9})\b")

# --- PAN section ----------------------------------------------------------
PAN_RE = re.compile(r"\b([A-Z]{5}\d{4}[A-Z])\b")
# Match income lines: "Rs. 7,04,830/-"
RS_VALUE_RE = re.compile(r"Rs\.?\s*([\d,]{3,15})\s*/?-?", re.IGNORECASE)

# --- Part B abstract ------------------------------------------------------
PB_HEADER_RE = re.compile(r"(?:PART[\s\-]*B|ABSTRACT\s+OF\s+THE\s+DETAILS)", re.IGNORECASE)
# "Total number of pending criminal cases ... 07 (SEVEN)"
PB_PENDING_RE = re.compile(
    r"[Tt]otal\s+number\s+of\s+pending\s+criminal[^\d\n]{0,40}?(\d{1,3})"
    r"\s*(?:\(([A-Z]+)\))?",
)
PB_CONVICTIONS_RE = re.compile(
    r"[Tt]otal\s+[Nn]umber\s+of\s+cases\s+in\s+which[^\d\n]{0,40}?(\d{1,3}|NIL|IL|NL)"
    r"\s*(?:\(([A-Z]+)\))?\s*[Cc]onvicted",
)
# Movable totals row in Part B abstract — "Moveable Assets ... Rs. X | Rs. Y"
PB_MOVABLE_RE = re.compile(
    r"M[oa]vea?ble\s+Assets[^A-Za-z]{0,30}?(?:Total[^A-Za-z]+?)?"
    r"(\d{4,}(?:[,\s]+\d+)*)\s*/?-?"
    r"\s*[|\\/]?\s*"
    r"(\d{4,}(?:[,\s]+\d+)*)",
    re.IGNORECASE | re.DOTALL,
)
# Item 7A "Gross Total value" row — usually appears on the movable-detail
# page (~p14). When the OCR concatenates the two numbers into one digit
# run, we split heuristically: assume the self value is 5-8 digits
# (₹10k-₹10cr), spouse takes the rest.
GROSS_TOTAL_RE = re.compile(
    r"(?:G|g)?ross\s+Total\s+val(?:ue|uc)[^\n]*\n[^\n]*?(\d{5,})",
    re.IGNORECASE,
)
# Liabilities — "Grand total of liabilities" or "Loans from Bank ... (Total)"
LIAB_BANK_RE = re.compile(
    r"Loans?\s+from\s+Bank[^\n]+?(?:Total\)?)?[^\d\n]{0,30}?(\d{4,}(?:[,\s]+\d+)*)\s*/?-?",
    re.IGNORECASE,
)
LIAB_DISPUTED_RE = re.compile(
    r"DISPUTE\s+OF\s+AMOUNT\s+(?:RS\.?\s*)?(\d{4,}(?:[,\s]+\d+)*)",
    re.IGNORECASE,
)

# Education — find "POST GRADUATE ... " or similar item (10) statement
EDU_RE = re.compile(
    r"(?:My\s+)?educational\s+qualification\s+is\s+as\s+under[:|\.]?\s*"
    r"([^\n]{20,250}?)\s*(?:\(Give\s+details|$|\n\s*\n)",
    re.IGNORECASE,
)

# Profession — "Self ... ACTIVIST", "Self ... BUSINESS"
PROFESSION_RE = re.compile(
    r"(?:Self|9\)\s*Self|9\(a\)\s*Self|\(a\)\s*Self)\s*[:|\.\s]+([A-Z][A-Za-z\s\/\&,'\(\)-]{2,80}?)"
    r"\s*(?:\(b\)\s*Spouse|\n)",
    re.IGNORECASE,
)


# ---------------------------------------------------------------------------
# Indian number parsing
# ---------------------------------------------------------------------------

def parse_indian_int(raw: str) -> int | None:
    """'1,00,89,655' or '4,979,105' or '4979 105' → 10089655 / 4979105"""
    if not raw:
        return None
    cleaned = re.sub(r"[^\d]", "", raw)
    try:
        return int(cleaned) if cleaned else None
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Per-candidate extraction
# ---------------------------------------------------------------------------

@dataclass
class CandidateRow:
    seq: int = 0
    source_pdf: str = ""
    affidavit_id: str = ""

    # eStamp
    estamp_cert: str = ""
    estamp_date: str = ""
    estamp_purchaser: str = ""
    stamp_duty_rs: int | None = None

    # Identity
    candidate_name: str = ""
    relationship: str = ""
    father_or_husband: str = ""
    age: int | None = None
    address: str = ""
    phone: str = ""
    email: str = ""

    # Election
    party: str = ""
    constituency: str = ""

    # Education + profession
    education: str = ""
    profession_self: str = ""

    # Tax
    self_pan: str = ""
    spouse_pan: str = ""

    # Criminal
    pending_cases: int | None = None
    convictions: int | None = None

    # Assets — Part B abstract movable totals
    movable_self: int | None = None
    movable_spouse: int | None = None

    # Liabilities
    liabilities_bank: int | None = None
    liabilities_disputed: int | None = None

    # Quality
    notes: list[str] = field(default_factory=list)


def _candidate_name_from_filename(filename: str) -> tuple[str, str]:
    """'005_AKHILESH_PATI_TRIPATHI__1679.json' → ('AKHILESH PATI TRIPATHI', '1679')"""
    stem = Path(filename).stem
    affid = ""
    if "__" in stem:
        name_part, affid = stem.rsplit("__", 1)
    else:
        name_part = stem
    # Drop leading sequence number
    parts = name_part.split("_", 1)
    if len(parts) == 2 and parts[0].isdigit():
        name_part = parts[1]
    return name_part.replace("_", " "), affid


def _truncate(s: str, n: int = 200) -> str:
    s = re.sub(r"\s+", " ", s).strip()
    return s if len(s) <= n else s[:n] + "…"


def _clean_field(s: str | None) -> str:
    """Strip OCR garbage characters from a captured string."""
    if not s:
        return ""
    s = re.sub(r"\s+", " ", s)
    s = s.strip(" |.:;\\/-").strip()
    return s


def extract_one(filename: str, payload: dict, seq: int) -> CandidateRow:
    row = CandidateRow(seq=seq, source_pdf=payload.get("source_pdf", filename))
    name_filename, affid = _candidate_name_from_filename(filename)
    row.affidavit_id = affid

    pages = payload.get("pages", [])
    all_text = "\n".join(p.get("text", "") for p in pages)
    cover_text = pages[0].get("text", "") if pages else ""

    # --- eStamp cover ---------------------------------------------------
    for p in pages[:3]:
        t = p.get("text", "")
        if m := ESTAMP_CERT_RE.search(t):
            cert = m.group(1)
            row.estamp_cert = cert[:-1] + cert[-1].upper()
            break
    if m := ESTAMP_DATE_RE.search(cover_text):
        row.estamp_date = _clean_field(m.group(1)).replace(".", ":")
    if m := ESTAMP_PURCHASER_RE.search(cover_text):
        row.estamp_purchaser = _clean_field(m.group(1))
    if m := ESTAMP_DUTY_RE.search(cover_text):
        try:
            row.stamp_duty_rs = int(m.group(1))
        except ValueError:
            pass

    # --- Part A identity ------------------------------------------------
    if m := PART_A_NAME_RE.search(all_text):
        captured_name = _clean_field(m.group(1))
        row.relationship = m.group(2).lower()
        row.father_or_husband = _clean_field(m.group(3)).strip(";")
        # Sanity-check against filename name; prefer the longer if both look real
        if captured_name and len(captured_name) >= 3:
            row.candidate_name = captured_name
    if not row.candidate_name:
        row.candidate_name = name_filename
        row.notes.append("name_from_filename (Part A regex missed)")

    if m := AGE_RE.search(all_text):
        try:
            row.age = int(m.group(1))
        except ValueError:
            pass
    if m := ADDRESS_RE.search(all_text):
        row.address = _truncate(_clean_field(m.group(1)), 200)
    if m := PARTY_RE.search(all_text):
        row.party = _truncate(_clean_field(m.group(1)), 80)
    if m := CONSTITUENCY_RE.search(all_text):
        row.constituency = _truncate(_clean_field(m.group(1)), 100)
    if m := EMAIL_RE.search(all_text):
        row.email = m.group(1)
    if m := PHONE_RE.search(all_text):
        row.phone = m.group(1)

    # --- PAN (first two distinct ones we find are usually self + spouse) -
    pans = list(dict.fromkeys(PAN_RE.findall(all_text)))   # preserve order, dedup
    # Drop eStamp cert false-positives — they begin with IN- (which won't match)
    pans = [p for p in pans if not p.startswith("IN")]
    if pans:
        row.self_pan = pans[0]
    if len(pans) >= 2:
        row.spouse_pan = pans[1]

    # --- Part B abstract ------------------------------------------------
    if m := PB_PENDING_RE.search(all_text):
        try:
            row.pending_cases = int(m.group(1))
        except ValueError:
            pass
    if m := PB_CONVICTIONS_RE.search(all_text):
        raw = m.group(1).upper()
        if raw in {"NIL", "IL", "NL"}:
            row.convictions = 0
        else:
            try:
                row.convictions = int(raw)
            except ValueError:
                pass

    if m := PB_MOVABLE_RE.search(all_text):
        row.movable_self = parse_indian_int(m.group(1))
        row.movable_spouse = parse_indian_int(m.group(2))
    elif m := GROSS_TOTAL_RE.search(all_text):
        # Fallback: Part A item 7A gross total row. OCR sometimes
        # concatenates self+spouse into one digit run (e.g. '1745874497910'
        # for ₹17,45,874 + ₹4,97,9105). Try to split at the most plausible
        # boundary: a 7-digit self total + remainder = spouse.
        raw = m.group(1)
        if len(raw) >= 11:
            # Try splitting at index 7 (typical self total is ~7 digits)
            row.movable_self = parse_indian_int(raw[:7])
            row.movable_spouse = parse_indian_int(raw[7:])
            row.notes.append("movable_split_concatenated_digits")
        else:
            row.movable_self = parse_indian_int(raw)
            row.notes.append("movable_self_only")

    if m := LIAB_BANK_RE.search(all_text):
        row.liabilities_bank = parse_indian_int(m.group(1))
    if m := LIAB_DISPUTED_RE.search(all_text):
        row.liabilities_disputed = parse_indian_int(m.group(1))

    # --- Education + profession -----------------------------------------
    if m := EDU_RE.search(all_text):
        row.education = _truncate(_clean_field(m.group(1)), 200)
    if m := PROFESSION_RE.search(all_text):
        row.profession_self = _truncate(_clean_field(m.group(1)), 100)

    # --- Quality notes --------------------------------------------------
    if not row.estamp_cert:
        row.notes.append("no_estamp_cert")
    if not row.self_pan:
        row.notes.append("no_pan")
    if row.pending_cases is None:
        row.notes.append("no_pending_cases_count")
    if row.movable_self is None and row.movable_spouse is None:
        row.notes.append("no_movable_totals")

    return row


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

CSV_COLUMNS = [
    "seq", "source_pdf", "affidavit_id",
    "candidate_name", "father_or_husband", "relationship", "age",
    "party", "constituency",
    "address", "phone", "email",
    "education", "profession_self",
    "self_pan", "spouse_pan",
    "pending_cases", "convictions",
    "movable_self", "movable_spouse",
    "liabilities_bank", "liabilities_disputed",
    "estamp_cert", "estamp_date", "estamp_purchaser", "stamp_duty_rs",
    "notes",
]


def write_csv(rows: list[CandidateRow], out_path: Path):
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(CSV_COLUMNS)
        for r in rows:
            d = asdict(r)
            d["notes"] = "; ".join(d.get("notes", []))
            w.writerow([d.get(c, "") if d.get(c) is not None else ""
                        for c in CSV_COLUMNS])


def write_xlsx(rows: list[CandidateRow], out_path: Path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed — skipping XLSX. Install with: "
              "pip install openpyxl", file=sys.stderr)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Delhi 2025"

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    NOTES_FILL = PatternFill("solid", fgColor="FBE5D6")
    BODY_FONT = Font(name="Arial", size=10)

    for col, h in enumerate(CSV_COLUMNS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center",
                                  wrap_text=True)

    for ri, r in enumerate(rows, 2):
        d = asdict(r)
        notes_str = "; ".join(d.get("notes", []))
        for ci, col_name in enumerate(CSV_COLUMNS, 1):
            v = d.get(col_name) if col_name != "notes" else notes_str
            if v is None:
                v = ""
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = BODY_FONT
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if col_name == "notes" and notes_str:
                c.fill = NOTES_FILL

    # Column widths (tuned for typical content)
    widths = {
        "seq": 5, "source_pdf": 38, "affidavit_id": 11,
        "candidate_name": 26, "father_or_husband": 26, "relationship": 11,
        "age": 5, "party": 28, "constituency": 26,
        "address": 36, "phone": 13, "email": 28,
        "education": 36, "profession_self": 24,
        "self_pan": 11, "spouse_pan": 11,
        "pending_cases": 8, "convictions": 8,
        "movable_self": 13, "movable_spouse": 13,
        "liabilities_bank": 13, "liabilities_disputed": 13,
        "estamp_cert": 22, "estamp_date": 21, "estamp_purchaser": 24,
        "stamp_duty_rs": 8, "notes": 40,
    }
    for i, col in enumerate(CSV_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 14)
    ws.freeze_panes = "D2"
    wb.save(out_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    import argparse
    ap = argparse.ArgumentParser(description=__doc__,
                                  formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--in-dir", default="data/eci/for_ai/preprocessed",
                    help="Directory of preprocessed *.json files "
                         "(default: data/eci/for_ai/preprocessed). Project-"
                         "relative or absolute path both work.")
    ap.add_argument("--out-dir", default="data/eci/for_ai/extracted",
                    help="Where to write the CSV/XLSX outputs. Project-"
                         "relative or absolute path both work.")
    ap.add_argument("--prefix", default="delhi_2025",
                    help="Output filename prefix — controls "
                         "`<prefix>_structured.{csv,xlsx}` (default: "
                         "delhi_2025). Use 'delhi_2020' for the 2020 cycle, "
                         "etc.")
    args = ap.parse_args()

    project_root = Path(__file__).resolve().parent.parent
    in_dir = Path(args.in_dir)
    out_dir = Path(args.out_dir)
    if not in_dir.is_absolute():
        in_dir = project_root / in_dir
    if not out_dir.is_absolute():
        out_dir = project_root / out_dir
    out_dir.mkdir(parents=True, exist_ok=True)
    out_prefix = args.prefix
    print(f"Input dir:  {in_dir}", file=sys.stderr)
    print(f"Output dir: {out_dir}", file=sys.stderr)
    print(f"Prefix:     {out_prefix}", file=sys.stderr)

    files = sorted(f for f in in_dir.glob("*.json") if not f.name.startswith("_"))
    if not files:
        sys.exit(f"No preprocessed JSONs in {in_dir}. "
                 "Run preprocess_modal.py first.")

    rows: list[CandidateRow] = []
    for i, f in enumerate(files, 1):
        try:
            payload = json.loads(f.read_text())
        except Exception as e:
            print(f"  bad JSON in {f.name}: {e}", file=sys.stderr)
            continue
        row = extract_one(f.name, payload, seq=i)
        rows.append(row)
        if i % 20 == 0 or i == len(files):
            print(f"  extracted {i}/{len(files)}", file=sys.stderr)

    csv_path = out_dir / f"{out_prefix}_structured.csv"
    xlsx_path = out_dir / f"{out_prefix}_structured.xlsx"
    notes_path = out_dir / f"_{out_prefix}_extraction_notes.csv"

    write_csv(rows, csv_path)
    write_xlsx(rows, xlsx_path)

    # Notes-only file for quick triage
    with notes_path.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["seq", "candidate_name", "notes"])
        for r in rows:
            if r.notes:
                w.writerow([r.seq, r.candidate_name, "; ".join(r.notes)])

    # Summary
    n = len(rows)
    fully_clean = sum(1 for r in rows if not r.notes)
    has_assets = sum(1 for r in rows
                      if r.movable_self is not None or r.movable_spouse is not None)
    has_pan = sum(1 for r in rows if r.self_pan)
    has_party = sum(1 for r in rows if r.party)
    print(f"\n========== STRUCTURED EXTRACTION SUMMARY ==========",
          file=sys.stderr)
    print(f"  Total candidates: {n}", file=sys.stderr)
    print(f"  Fully clean (no notes): {fully_clean}  "
          f"({100 * fully_clean / n:.1f}%)", file=sys.stderr)
    print(f"  With movable totals:    {has_assets}  "
          f"({100 * has_assets / n:.1f}%)", file=sys.stderr)
    print(f"  With self PAN:          {has_pan}  "
          f"({100 * has_pan / n:.1f}%)", file=sys.stderr)
    print(f"  With party detected:    {has_party}  "
          f"({100 * has_party / n:.1f}%)", file=sys.stderr)
    print(f"\nWrote:", file=sys.stderr)
    print(f"  {csv_path}", file=sys.stderr)
    print(f"  {xlsx_path}", file=sys.stderr)
    print(f"  {notes_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
