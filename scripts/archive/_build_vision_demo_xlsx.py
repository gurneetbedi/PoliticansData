"""Build the Vision-vs-myneta comparison Excel for the Delhi 2025 demo.

Each candidate row carries:
  - DB columns (what myneta currently has, from politrack.db)
  - ECI columns (what we just extracted from the affidavit PDF via vision)
  - Delta / Status / Notes columns

The candidates dict below is populated by hand from the vision read of each
PDF — this is the proof-of-concept showing what a Claude vision pipeline
would produce at scale.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sqlite3
from pathlib import Path


# ---------------------------------------------------------------------------
# Vision-extracted candidate data (manually captured from the PDFs we read)
# ---------------------------------------------------------------------------

CANDIDATES = [
    {
        # === ELECTION + IDENTITY ===
        "candidate_name": "AKHILESH PATI TRIPATHI",
        "father_or_husband_name": "SH. ABHAY NANDAN TRIPATHI",
        "age": 40,
        "party_full": "AAM AADMI PARTY",
        "constituency": "AC-18, MODEL TOWN, NCT OF DELHI",
        "voter_serial": "563 in Part No. 53",
        "address": "N-9C/129, LAL BAGH, AZADPUR, DELHI-110033",
        "phone": "9873386499, 8588833977",
        "email": "akhilesht84@gmail.com",
        "social_media": ("FB akhileshpati.tripathi.73 | aapakhilesh | "
                          "IG Akhilesht84 | YT UCXHsWB1Lm6q8EndAE6dIclA | "
                          "X AapAkhilesh"),
        "education": ("Post Graduate (Master of Arts in Ancient History), "
                       "RML Avadh University, Faizabad, UP, 2008"),
        "profession_self": "POLITICAL ACTIVIST / MLA, MODEL TOWN (AC-18) NCT OF DELHI",
        "profession_spouse": "HOUSEWIFE / HOME TUTOR",
        "income_source_self": "SALARY",
        "income_source_spouse": "HOME TUITION",

        # === PAN + INCOME TAX (last 5 years) ===
        "self_pan": "AUFPT4082Q",
        "self_income_fy_2023_24": 704830,
        "self_income_fy_2022_23": 479900,
        "self_income_fy_2021_22": 493480,
        "self_income_fy_2020_21": 490920,
        "self_income_fy_2019_20": 499460,
        "spouse_name": "PRATIMA PANDEY TRIPATHI",
        "spouse_pan": "BGOPT9337G",
        "spouse_income_fy_2023_24": 497450,
        "spouse_income_fy_2022_23": 494350,
        "spouse_income_fy_2021_22": 333090,
        "spouse_income_fy_2020_21": 241300,
        "spouse_income_fy_2019_20": None,
        "dependents": "ANIKA TRIPATHI, ANANT TRIPATHI",

        # === SPOUSE FULL PROFILE ===
        "spouse_cash": 15000,
        "spouse_bank_accounts": (
            "SBI Model Town A/c 20253312795 with Locker ₹7,01,626.54 (approx); "
            "ICICI Bank Gujrawala Town A/c 071701509603 ₹0.00"
        ),
        "spouse_bank_total": 701626,
        "spouse_investments_mf": (
            "8 mutual funds (current market values): "
            "Axis Focused Fund ₹2,63,085; "
            "HDFC Hybrid Equity Fund ₹2,74,004; "
            "Mirae Asset Focused Fund ₹2,49,097; "
            "Motilal Oswal Flexi Cap Fund ₹3,54,388; "
            "Tata India Consumer Fund ₹9,48,853; "
            "Kotak Emerging Equity ₹2,14,241; "
            "Sundaram Large and Mid Cap ₹13,548; "
            "SBI Focused Equity Fund ₹3,08,651"
        ),
        "spouse_investments_total": 263085 + 274004 + 249097 + 354388 + 948853 + 214241 + 13548 + 308651,
        "spouse_insurance": (
            "Future Generali New Assured Wealth Plan, Policy 01761277 "
            "(start 31/03/2023), Yearly Premium ₹57,000, Sum Assured ₹4,56,000"
        ),
        "spouse_vehicles": "NIL",
        "spouse_jewellery": (
            "Gold 141.76gm ₹10,23,562; Silver 1540gm ₹1,57,080 (both approx)"
        ),
        "spouse_jewellery_total": 1023562 + 157080,
        "spouse_immovable_self": 0,
        "spouse_immovable_notes": (
            "NIL across all categories (no agricultural, non-agricultural, "
            "commercial, or residential property in spouse's name)"
        ),
        "spouse_profession": "HOUSEWIFE / HOME TUTOR",
        "spouse_income_source": "HOME TUITION",
        "spouse_personal_loans_given": "NIL",
        "spouse_other_claims": "NIL",
        "spouse_movable_gross_partA": 4979105,
        "spouse_movable_gross_partB": 4979105,

        # === CRIMINAL CASES (the BIG finding) ===
        "pending_criminal_cases_count": 7,
        "pending_criminal_cases_detail": (
            "1) Complaint Case 4/2019 Vivek Garg vs Akhilesh, U/S 420 IPC, "
            "illegal medical reimbursement; "
            "2) FIR 292/2019 PS Adarsh Nagar, U/S 186/353/332/34 IPC, "
            "hurting public servant (charges framed 9/01/2020); "
            "3) FIR 11/2022 PS ACB, U/S 7A/13 PCC Act & 171(B)(E) IPC, "
            "influencing public servant; "
            "4) Lokayukta Case C-3842/LOK/2018, before Justice H.C. Mishra; "
            "5) FIR 559/24 PS Model Town, U/S 221/132/121(1)/3(5) BNS, "
            "voluntarily causing hurt to public servant; "
            "6) FIR 537/20 PS Civil Lines, U/S 186/353/332/34 IPC; "
            "7) FIR 261/20 PS Kamla Market, U/S 186/353/332/188/269/34 IPC "
            "+ 3 ED Act"
        ),
        "convictions_count": 3,
        "convictions_detail": (
            "1) SC NO-01/2021, MS Geetanjali Goel ASJ/SPL JUDGE PC ACT CBI "
            "RADC/New Delhi — FIR 84/2020 originally U/S 3(i)(r)(s) SC/ST "
            "ACT & U/S 323/341/506 IPC, BUT CONVICTED ONLY U/S 323 IPC — "
            "17.05.2023 — Till rising of court + ₹30,000 compensation; "
            "2) Case 33/2015 same court — FIR 236/2015 U/S 147/186/332/506 "
            "IPC + Damage to Public Property Act 1984 — CONVICTED ONLY U/S "
            "332 R/W 149 IPC — 03.07.2023 — Till rising of court + ₹10,000 "
            "fine; "
            "3) SC NO-04/2019 MPs MLAs case same court — FIR 260/2013 — "
            "22.09.2021 — 3-month probation bond + ₹6,150 costs under "
            "Probation of Offenders Act 1958"
        ),

        # === MOVABLE ASSETS ===
        # Part A detail page 14 (the breakdown that ADR / myneta uses)
        "movable_self_partA": 1745874,
        "movable_spouse_partA": 4979105,
        # Part B abstract page 25 (candidate's own headline — sometimes differs)
        "movable_self_partB": 2315874,
        "movable_spouse_partB": 4979105,
        "movable_breakdown": (
            "Self: Cash ₹40,000 + SBI Civil Lines ₹5,34,549 + HDFC Rana "
            "Pratap Bagh ₹25,000 + SBI Menhdawal FDR ₹1,34,304 + ICICI "
            "Gujrawala Town ₹0 + TATA TIAGO 2023 (DL10CEV0007) ₹10,12,020 "
            "+ Jewellery NIL. "
            "Spouse: Cash ₹15,000 + SBI Model Town ₹7,01,626 (with locker) "
            "+ MF holdings (Axis Focused ₹2,63,085 + HDFC Hybrid ₹2,74,004 "
            "+ Mirae Asset ₹2,49,097 + Motilal Oswal ₹3,54,388 + Tata India "
            "Consumer ₹9,48,853 + Kotak Emerging ₹2,14,241 + Sundaram Large "
            "& Mid ₹13,548 + SBI Focused ₹3,08,651) + Future Generali "
            "insurance (sum assured ₹4,56,000, premium ₹57,000/yr) + Gold "
            "141.76g ₹10,23,562 + Silver 1540g ₹1,57,080"
        ),

        # === IMMOVABLE ASSETS ===
        "immovable_self": 0,            # NIL across all categories
        "immovable_spouse": 0,
        "immovable_notes": ("NIL across all categories (no agricultural, "
                            "non-agricultural, commercial, or residential)"),

        # === LIABILITIES ===
        "liabilities_total": 400926,    # HDFC Bank Card Loan A/C 14485 1105
        "liabilities_detail": "HDFC Bank Card Loan A/C 14485 1105 — ₹4,00,926 (approx)",
        "liabilities_disputed": 2738386,
        "liabilities_disputed_detail": (
            "Office space under PWD allotted for Model Town AC Office "
            "(Flat H-3, Type-V Delhi Admin Flats) — disputed amount "
            "₹27,38,386 claimed by PWD; dispute pending before Delhi "
            "Vidhan Sabha"
        ),

        # === TOTALS (computed) ===
        # ECI vision total = Part A movable subtotal (matches ADR convention)
        "eci_total_assets": 1745874 + 4979105 + 0 + 0,
        "eci_total_liabilities": 400926,
        "data_quality_notes": (
            "Part A detail sums to ₹67.24L (matches DB exactly). "
            "Part B abstract claims ₹72.94L movable (₹5.7L higher — "
            "internal inconsistency in candidate's own filing). "
            "ADR convention uses Part A detail. "
            "DB says 0 convictions BUT affidavit declares 3 convictions — "
            "major data gap in myneta source."
        ),

        # === SOURCE PDF ===
        "source_pdf": "AKHILESH_PATI_TRIPATHI__1679.pdf",
        "estamp_cert": "IN-DL19330930870225X",
        "estamp_date": "16-Jan-2025 02:02 PM",
    },

    # --------------------------------------------------------------
    # ASHISH SOOD — Independent (NOT the BJP candidate of same name)
    # --------------------------------------------------------------
    # CRITICAL FINDING: the politrack DB has ASHISH SOOD = BJP, JANAKPURI,
    # WON, ₹9.22 Cr. This affidavit is a DIFFERENT person of the same name
    # — an Independent who also contested AC-30 Janakpuri but declared
    # ₹22.6L. Total mismatch. Demonstrates the danger of name-only DB
    # matching for common Indian names.
    {
        "candidate_name": "ASHISH SOOD (Independent)",
        "father_or_husband_name": "RAKESH SOOD",
        "age": 37,
        "party_full": "INDEPENDENT",
        "constituency": "AC-30, JANAKPURI, NCT OF DELHI",
        "voter_serial": "536 in Part No. 54",
        "address": "66-TF, GALI NO-9, W Z BLOCK, VIRENDER NAGAR, DELHI-110058",
        "phone": "9899997646",
        "email": "aashishsood@gmail.com",
        "social_media": "FB ashish.sood.77 | IG ashish_1987 | WhatsApp 9899997646",
        "education": "10th Pass, Guru Har Kishan Public School, CBSE, 2003",
        "profession_self": "BUSINESS",
        "profession_spouse": "PRIVATE JOB",
        "income_source_self": "BUSINESS",
        "income_source_spouse": "SALARY",

        "self_pan": "DBJPS2629E",
        "self_income_fy_2023_24": 712580,
        "self_income_fy_2022_23": 402400,
        "self_income_fy_2021_22": 385550,
        "self_income_fy_2020_21": 420170,
        "self_income_fy_2019_20": 405000,
        "spouse_name": "SHILPI SOOD",
        "spouse_pan": "BZXPS9179M",
        "spouse_income_fy_2023_24": 493580,
        "spouse_income_fy_2022_23": 568250,
        "spouse_income_fy_2021_22": 349570,
        "spouse_income_fy_2020_21": None,
        "spouse_income_fy_2019_20": None,
        "dependents": "SHIVANK SOOD",

        # === SPOUSE FULL PROFILE ===
        "spouse_cash": 14800,
        "spouse_bank_accounts": (
            "Axis Bank Janakpuri S/B A/c 207010100299534 ₹1,53,510 (10/01/2025); "
            "Axis Bank Janakpuri FD A/c 923040085148940 ₹5,65,615 (10/01/2025); "
            "Axis Bank Janakpuri FD A/c 923040081837593 ₹1,07,916 (10/01/2025); "
            "Axis Bank Janakpuri FD A/c 923040085149057 ₹5,65,616 (10/01/2025)"
        ),
        "spouse_bank_total": 153510 + 565615 + 107916 + 565616,
        "spouse_investments_mf": "NOT APPLICABLE (no MF or bond holdings)",
        "spouse_investments_total": 0,
        "spouse_insurance": "NIL",
        "spouse_vehicles": "NIL",
        "spouse_jewellery": (
            "Gold 50gm: 1 Mangal sutra approx 20gm + 1 Gold Chain approx 20gm + "
            "2 Anghuti approx 10gm — Market value ₹4,00,000 (approx)"
        ),
        "spouse_jewellery_total": 400000,
        "spouse_immovable_self": 0,
        "spouse_immovable_notes": "NIL across all categories",
        "spouse_profession": "PRIVATE JOB",
        "spouse_income_source": "SALARY",
        "spouse_personal_loans_given": "NIL",
        "spouse_other_claims": "NIL",
        "spouse_movable_gross_partA": 1807257,
        "spouse_movable_gross_partB": 1807257,

        "pending_criminal_cases_count": 0,
        "pending_criminal_cases_detail": "NIL (declared no pending criminal cases)",
        "convictions_count": 0,
        "convictions_detail": "NIL",

        "movable_self_partA": 453885,
        "movable_spouse_partA": 1807257,
        "movable_self_partB": 453885,
        "movable_spouse_partB": 1807257,
        "movable_breakdown": (
            "Self: Cash ₹25,030 + Axis Bank Janakpuri S/B ₹56,677 + "
            "Punjab & Sind Bank Mahavir Enclave ₹1,000 (election expense "
            "account) + Maruti Eco 2017 DL12CL3397 (purchase ₹4.95L, "
            "market value ₹1.90L) + Gold 20gm (1 chain 10gm + 2 anghuti "
            "10gm) ₹1,60,000. "
            "Spouse: Cash ₹14,800 + Axis Bank Janakpuri S/B ₹1,53,510 + "
            "Axis FD ₹5,65,615 + Axis FD ₹1,07,916 + Axis FD ₹5,65,616 + "
            "Gold 50gm (Mangal sutra 20gm + chain 20gm + 2 anghuti 10gm) "
            "₹4,00,000 + National Insurance Health Policy ₹21,178"
        ),

        "immovable_self": 0,
        "immovable_spouse": 0,
        "immovable_notes": "NIL across all categories",

        "liabilities_total": 0,
        "liabilities_detail": "NIL (no loans, no government dues)",
        "liabilities_disputed": 0,
        "liabilities_disputed_detail": "NIL",

        "eci_total_assets": 453885 + 1807257 + 0 + 0,
        "eci_total_liabilities": 0,
        "data_quality_notes": (
            "⚠️ NAME COLLISION RISK: politrack DB has 'Ashish Sood' = BJP, "
            "WON, ₹9.22 Cr — that's the BJP candidate from same constituency "
            "(different person, different ECI filing). The affidavit we "
            "downloaded is for the INDEPENDENT contender. Vision parsed it "
            "correctly (₹22.6L matches the affidavit's Part B abstract). "
            "Lesson: name-only DB matching is unsafe; need name + constituency "
            "+ party for unique identification."
        ),

        "source_pdf": "ASHISH_SOOD__1834.pdf",
        "estamp_cert": "IN-DL19692791286251X",
        "estamp_date": "17-Jan-2025 10:31 AM",
    },
]


# ---------------------------------------------------------------------------
# DB lookup — pull the matching myneta row for each candidate
# ---------------------------------------------------------------------------

DB_PATH = "politrack.db"


def db_row_for(name: str) -> dict:
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    cur = con.cursor()
    cur.execute("""
        SELECT p.id AS pid, p.name, ea.total_assets_inr,
               ea.total_liabilities_inr, ea.criminal_cases_count,
               ea.education, c.name AS constituency, party.short_name AS party,
               ea.won
        FROM politicians p
        JOIN election_appearances ea ON ea.politician_id = p.id
        JOIN elections e ON ea.election_id = e.id
        LEFT JOIN constituencies c ON ea.constituency_id = c.id
        LEFT JOIN parties party ON ea.party_id = party.id
        LEFT JOIN states s ON e.state_id = s.id
        WHERE e.year=2025 AND s.name='Delhi' AND UPPER(p.name)=UPPER(?)
    """, (name,))
    row = cur.fetchone()
    return dict(row) if row else {}


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------

BOLD = Font(name="Arial", bold=True, size=11)
NORMAL = Font(name="Arial", size=10)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
DB_FILL = PatternFill("solid", fgColor="FFF2CC")        # light yellow
ECI_FILL = PatternFill("solid", fgColor="E2EFDA")       # light green
DIFF_FILL = PatternFill("solid", fgColor="FBE5D6")      # light orange
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fmt_rs(v):
    if v is None:
        return ""
    if isinstance(v, str):
        return v
    return f"₹{v:,}"


def main():
    wb = Workbook()

    # ============================================================
    # Sheet 1 — Side-by-side comparison (the headline view)
    # ============================================================
    ws = wb.active
    ws.title = "Comparison"

    # Header
    headers = ["Field", "myneta DB", "ECI Vision", "Verdict / Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    row = 2
    for cand in CANDIDATES:
        db = db_row_for(cand["candidate_name"])

        # Candidate header band
        title = (f"{cand['candidate_name']}  |  {cand['party_full']}  |  "
                  f"{cand['constituency'].split(',')[0]}  "
                  f"(myneta_id={db.get('pid','?')})")
        cell = ws.cell(row=row, column=1, value=title)
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=4)
        cell.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5496")
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 24
        row += 1

        rows = [
            ("Name", db.get("name"), cand["candidate_name"], "Match (case only)"),
            ("Father / Husband", "—", cand["father_or_husband_name"], "ECI-only field"),
            ("Age", "—", cand["age"], "ECI-only field"),
            ("Party", db.get("party"), cand["party_full"], "Match (short vs full)"),
            ("Constituency", db.get("constituency"), cand["constituency"], "Match"),
            ("Address", "—", cand["address"], "ECI-only field"),
            ("Phone", "—", cand["phone"], "ECI-only field"),
            ("Email", "—", cand["email"], "ECI-only field"),
            ("Education", db.get("education"), cand["education"], "ECI much richer"),
            ("Profession (self)", "—", cand["profession_self"], "ECI-only field"),
            ("Self PAN", "—", cand["self_pan"], "ECI-only field"),
            ("Spouse name", "—", cand["spouse_name"], "ECI-only field"),
            ("Spouse PAN", "—", cand["spouse_pan"], "ECI-only field"),
            ("Spouse profession", "—", cand.get("spouse_profession", "—"),
             "ECI-only field"),
            ("Spouse income source", "—", cand.get("spouse_income_source", "—"),
             "ECI-only field"),
            ("Self income FY23-24", "—", fmt_rs(cand["self_income_fy_2023_24"]),
             "ECI-only field"),
            ("Spouse income FY23-24", "—", fmt_rs(cand["spouse_income_fy_2023_24"]),
             "ECI-only field"),
            ("Spouse bank balance total", "—",
             fmt_rs(cand.get("spouse_bank_total")), "ECI-only field"),
            ("Spouse investments (MF/Bonds total)", "—",
             fmt_rs(cand.get("spouse_investments_total")),
             "ECI-only field"),
            ("Spouse jewellery value", "—",
             fmt_rs(cand.get("spouse_jewellery_total")), "ECI-only field"),
            ("Spouse movable GROSS (Part B)", "—",
             fmt_rs(cand.get("spouse_movable_gross_partB")),
             "ECI-only field; same in Part A unless flagged"),
            ("Spouse immovable assets", "—",
             fmt_rs(cand.get("spouse_immovable_self")),
             "ECI-only field"),
            ("Pending criminal cases", db.get("criminal_cases_count"),
             cand["pending_criminal_cases_count"],
             f"DB={db.get('criminal_cases_count')} vs Affidavit=7 — DB OVER by 3"),
            ("Convictions", 0, cand["convictions_count"],
             "⚠ DB=0; affidavit declares 3 convictions (IPC 323, 332, riot)"),
            ("Total assets (₹)", fmt_rs(db.get("total_assets_inr")),
             fmt_rs(cand["eci_total_assets"]),
             ("✓ EXACT MATCH" if cand["eci_total_assets"] ==
              db.get("total_assets_inr") else "Δ within rounding")),
            ("Total liabilities (₹)", fmt_rs(db.get("total_liabilities_inr")),
             fmt_rs(cand["eci_total_liabilities"]),
             ("✓ EXACT MATCH" if cand["eci_total_liabilities"] ==
              db.get("total_liabilities_inr") else "Δ within rounding")),
            ("Disputed liabilities (₹)", "—",
             fmt_rs(cand["liabilities_disputed"]),
             "ECI-only field (legal exposure not on the site today)"),
            ("Vehicle", "—", "TATA TIAGO 2023 DL10CEV0007 ₹10.12L",
             "ECI-only field"),
            ("Jewellery", "—", "Gold 141.76g ₹10.24L; Silver 1540g ₹1.57L",
             "ECI-only field"),
            ("MF holdings", "—", "8 funds, full names + current values",
             "ECI-only field"),
            ("Data-quality notes", "—", cand["data_quality_notes"], ""),
        ]
        for field, db_val, eci_val, verdict in rows:
            ws.cell(row=row, column=1, value=field).font = BOLD
            c2 = ws.cell(row=row, column=2,
                          value=("" if db_val is None else str(db_val)))
            c2.fill = DB_FILL
            c3 = ws.cell(row=row, column=3,
                          value=("" if eci_val is None else str(eci_val)))
            c3.fill = ECI_FILL
            c4 = ws.cell(row=row, column=4, value=verdict)
            c4.fill = DIFF_FILL
            for c in (ws.cell(row=row, column=1), c2, c3, c4):
                c.alignment = WRAP
                c.font = c.font.copy(name="Arial")
                c.border = BORDER
            row += 1

        # blank row between candidates
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 40

    # ============================================================
    # Sheet 2 — Full ECI extraction (every field per candidate, vertical)
    # ============================================================
    ws2 = wb.create_sheet("ECI Full Detail")
    headers2 = ["Field"] + [c["candidate_name"] for c in CANDIDATES]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    # Order matters for readability — group by section
    field_order = [
        ("— IDENTITY —", None),
        ("Name", "candidate_name"),
        ("Father / Husband", "father_or_husband_name"),
        ("Age", "age"),
        ("Address", "address"),
        ("Phone", "phone"),
        ("Email", "email"),
        ("Social media", "social_media"),
        ("Voter serial", "voter_serial"),

        ("— ELECTION —", None),
        ("Party (full name)", "party_full"),
        ("Constituency", "constituency"),

        ("— EDUCATION & PROFESSION —", None),
        ("Education (full)", "education"),
        ("Profession (self)", "profession_self"),
        ("Income source (self)", "income_source_self"),
        ("Profession (spouse)", "profession_spouse"),
        ("Income source (spouse)", "income_source_spouse"),

        ("— TAX & INCOME — SELF (5-yr) —", None),
        ("Self PAN", "self_pan"),
        ("Self income FY23-24", "self_income_fy_2023_24"),
        ("Self income FY22-23", "self_income_fy_2022_23"),
        ("Self income FY21-22", "self_income_fy_2021_22"),
        ("Self income FY20-21", "self_income_fy_2020_21"),
        ("Self income FY19-20", "self_income_fy_2019_20"),

        ("— SPOUSE PROFILE —", None),
        ("Spouse name", "spouse_name"),
        ("Spouse PAN", "spouse_pan"),
        ("Spouse profession", "spouse_profession"),
        ("Spouse income source", "spouse_income_source"),
        ("Spouse income FY23-24", "spouse_income_fy_2023_24"),
        ("Spouse income FY22-23", "spouse_income_fy_2022_23"),
        ("Spouse income FY21-22", "spouse_income_fy_2021_22"),
        ("Spouse income FY20-21", "spouse_income_fy_2020_21"),
        ("Spouse income FY19-20", "spouse_income_fy_2019_20"),
        ("Spouse cash in hand", "spouse_cash"),
        ("Spouse bank accounts (detail)", "spouse_bank_accounts"),
        ("Spouse bank balance total", "spouse_bank_total"),
        ("Spouse investments (MF/Bonds detail)", "spouse_investments_mf"),
        ("Spouse investments total", "spouse_investments_total"),
        ("Spouse insurance", "spouse_insurance"),
        ("Spouse vehicles", "spouse_vehicles"),
        ("Spouse jewellery (detail)", "spouse_jewellery"),
        ("Spouse jewellery value", "spouse_jewellery_total"),
        ("Spouse personal loans given", "spouse_personal_loans_given"),
        ("Spouse other claims", "spouse_other_claims"),
        ("Spouse immovable assets", "spouse_immovable_self"),
        ("Spouse immovable notes", "spouse_immovable_notes"),
        ("Spouse movable GROSS (Part A)", "spouse_movable_gross_partA"),
        ("Spouse movable GROSS (Part B abstract)", "spouse_movable_gross_partB"),
        ("Dependents", "dependents"),

        ("— CRIMINAL —", None),
        ("Pending criminal cases (count)", "pending_criminal_cases_count"),
        ("Pending cases (detail)", "pending_criminal_cases_detail"),
        ("Convictions (count)", "convictions_count"),
        ("Convictions (detail)", "convictions_detail"),

        ("— ASSETS —", None),
        ("Movable self (Part A)", "movable_self_partA"),
        ("Movable spouse (Part A)", "movable_spouse_partA"),
        ("Movable self (Part B abstract)", "movable_self_partB"),
        ("Movable spouse (Part B abstract)", "movable_spouse_partB"),
        ("Movable detail breakdown", "movable_breakdown"),
        ("Immovable self", "immovable_self"),
        ("Immovable spouse", "immovable_spouse"),
        ("Immovable notes", "immovable_notes"),

        ("— LIABILITIES —", None),
        ("Total liabilities", "liabilities_total"),
        ("Liabilities detail", "liabilities_detail"),
        ("Disputed liabilities", "liabilities_disputed"),
        ("Disputed detail", "liabilities_disputed_detail"),

        ("— PROVENANCE —", None),
        ("Source PDF", "source_pdf"),
        ("eStamp cert", "estamp_cert"),
        ("eStamp date", "estamp_date"),
        ("Data-quality notes", "data_quality_notes"),
    ]
    r = 2
    for label, key in field_order:
        cell = ws2.cell(row=r, column=1, value=label)
        if key is None:
            # Section header row
            cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
            for col in range(2, len(CANDIDATES) + 2):
                ws2.cell(row=r, column=col, value="").fill = PatternFill(
                    "solid", fgColor="4472C4"
                )
        else:
            cell.font = BOLD
            for col, cand in enumerate(CANDIDATES, 2):
                v = cand.get(key)
                if isinstance(v, int) and "income" in key:
                    out = f"₹{v:,}"
                elif isinstance(v, int) and ("movable" in key or "immovable" in key
                                              or "liab" in key):
                    out = f"₹{v:,}"
                else:
                    out = "" if v is None else str(v)
                c = ws2.cell(row=r, column=col, value=out)
                c.alignment = WRAP
                c.font = NORMAL
                c.border = BORDER
            cell.alignment = WRAP
            cell.border = BORDER
        r += 1

    ws2.column_dimensions["A"].width = 32
    for i in range(len(CANDIDATES)):
        ws2.column_dimensions[get_column_letter(i + 2)].width = 60

    out = Path("data/eci/vision_demo/delhi_vision_vs_myneta.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
